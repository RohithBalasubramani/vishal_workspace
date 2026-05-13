"""PDF and Excel exporters for quotations."""

from __future__ import annotations

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from .calculator import compute_totals, format_inr

EXPORTS_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "data",
    "exports",
)
os.makedirs(EXPORTS_DIR, exist_ok=True)


def _safe(s) -> str:
    return "" if s is None else str(s)


def _export_path(quote_number: str, ext: str) -> str:
    safe_qn = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in (quote_number or "quote"))
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(EXPORTS_DIR, f"{safe_qn}_{stamp}.{ext}")


# ──────────────────────────────────────────────────────────────
# PDF
# ──────────────────────────────────────────────────────────────

def export_pdf(header: dict, items: list[dict]) -> str:
    """Render a quotation as PDF. Returns absolute path."""
    totals = compute_totals(items, gst_rate=header.get("gst_rate", 18))
    path = _export_path(header.get("quote_number", "quote"), "pdf")

    doc = SimpleDocTemplate(
        path,
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title=f"Quotation {_safe(header.get('quote_number'))}",
    )

    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Title"], fontSize=18, spaceAfter=6)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=11, spaceAfter=2)
    normal = styles["Normal"]
    small = ParagraphStyle("small", parent=normal, fontSize=8, leading=10)

    flow = []
    flow.append(Paragraph("QUOTATION", h1))

    # Header info — two-column table
    qd = _safe(header.get("quote_date") or "")
    vu = _safe(header.get("valid_until") or "")
    info_rows = [
        ["Quote #", _safe(header.get("quote_number")), "Date", qd],
        ["Customer", _safe(header.get("customer_name")), "Valid until", vu],
        ["Project", _safe(header.get("project_name")), "Status", _safe(header.get("status", "draft"))],
    ]
    addr = _safe(header.get("customer_address"))
    if addr:
        info_rows.append(["Address", Paragraph(addr.replace("\n", "<br/>"), small), "", ""])

    info_tbl = Table(info_rows, colWidths=[25 * mm, 75 * mm, 25 * mm, 55 * mm])
    info_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("LINEBELOW", (0, -1), (-1, -1), 0.25, colors.grey),
    ]))
    flow.append(info_tbl)
    flow.append(Spacer(1, 8))
    flow.append(Paragraph("Line Items", h2))

    # Items table
    head = ["#", "Type", "Model", "Description", "Brand", "Qty", "Unit Price", "Disc %", "Line Total"]
    data = [head]
    for idx, it in enumerate(totals["items"], start=1):
        data.append([
            str(idx),
            _safe(it.get("item_type", "")).capitalize(),
            _safe(it.get("source_model")),
            Paragraph(_safe(it.get("item_name")), small),
            _safe(it.get("brand")),
            f"{float(it.get('quantity', 0)):g}",
            format_inr(it.get("unit_price", 0)),
            f"{float(it.get('discount_pct', 0)):g}",
            format_inr(it.get("line_total", 0)),
        ])

    items_tbl = Table(
        data,
        colWidths=[8 * mm, 16 * mm, 28 * mm, 56 * mm, 20 * mm, 12 * mm, 24 * mm, 14 * mm, 26 * mm],
        repeatRows=1,
    )
    items_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f3a5f")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("ALIGN", (5, 1), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 1), (0, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f7fb")]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
    ]))
    flow.append(items_tbl)
    flow.append(Spacer(1, 8))

    # Totals
    totals_rows = [
        ["Subtotal", format_inr(totals["subtotal"])],
        [f"CGST ({totals['gst_rate']/2:g}%)", format_inr(totals["cgst"])],
        [f"SGST ({totals['gst_rate']/2:g}%)", format_inr(totals["sgst"])],
        ["Grand Total", format_inr(totals["grand_total"])],
    ]
    totals_tbl = Table(totals_rows, colWidths=[140 * mm, 40 * mm], hAlign="RIGHT")
    totals_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (0, 0), (0, -1), "RIGHT"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("LINEABOVE", (0, -1), (-1, -1), 0.5, colors.black),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, -1), (-1, -1), 12),
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.HexColor("#1f3a5f")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
    ]))
    flow.append(totals_tbl)

    if header.get("notes"):
        flow.append(Spacer(1, 10))
        flow.append(Paragraph("Notes", h2))
        flow.append(Paragraph(_safe(header.get("notes")).replace("\n", "<br/>"), small))

    flow.append(Spacer(1, 14))
    flow.append(Paragraph(
        "This is a system-generated quotation. Prices are indicative and subject to confirmation.",
        small,
    ))

    doc.build(flow)
    return path


# ──────────────────────────────────────────────────────────────
# Excel
# ──────────────────────────────────────────────────────────────

def export_excel(header: dict, items: list[dict]) -> str:
    """Render a quotation as Excel. Returns absolute path."""
    totals = compute_totals(items, gst_rate=header.get("gst_rate", 18))
    path = _export_path(header.get("quote_number", "quote"), "xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=16, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3A5F")
    band_fill = PatternFill("solid", fgColor="F5F7FB")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    right = Alignment(horizontal="right")
    center = Alignment(horizontal="center")
    wrap = Alignment(wrap_text=True, vertical="top")

    ws.merge_cells("A1:I1")
    ws["A1"] = "QUOTATION"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A1"].fill = header_fill
    ws.row_dimensions[1].height = 28

    row = 3
    info = [
        ("Quote #", header.get("quote_number"), "Date", header.get("quote_date")),
        ("Customer", header.get("customer_name"), "Valid until", header.get("valid_until")),
        ("Project", header.get("project_name"), "Status", header.get("status", "draft")),
        ("Address", header.get("customer_address"), "GST rate", f"{totals['gst_rate']:g}%"),
    ]
    for label1, value1, label2, value2 in info:
        ws.cell(row=row, column=1, value=label1).font = bold
        ws.cell(row=row, column=2, value=_safe(value1))
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row=row, column=5, value=label2).font = bold
        ws.cell(row=row, column=6, value=_safe(value2))
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=9)
        row += 1

    row += 1
    headers = ["#", "Type", "Model", "Description", "Brand", "Qty", "Unit Price", "Disc %", "Line Total"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = center
        c.border = border
    row += 1

    for idx, it in enumerate(totals["items"], start=1):
        values = [
            idx,
            _safe(it.get("item_type", "")).capitalize(),
            _safe(it.get("source_model")),
            _safe(it.get("item_name")),
            _safe(it.get("brand")),
            float(it.get("quantity", 0) or 0),
            float(it.get("unit_price", 0) or 0),
            float(it.get("discount_pct", 0) or 0),
            float(it.get("line_total", 0) or 0),
        ]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = border
            if idx % 2 == 0:
                c.fill = band_fill
            if col == 1:
                c.alignment = center
            elif col == 4:
                c.alignment = wrap
            elif col >= 6:
                c.alignment = right
            if col in (7, 9):
                c.number_format = '"₹"#,##0.00'
            elif col == 8:
                c.number_format = '0.00"%"'
        row += 1

    row += 1
    totals_rows = [
        ("Subtotal", totals["subtotal"]),
        (f"CGST ({totals['gst_rate']/2:g}%)", totals["cgst"]),
        (f"SGST ({totals['gst_rate']/2:g}%)", totals["sgst"]),
        ("Grand Total", totals["grand_total"]),
    ]
    for label, value in totals_rows:
        ws.cell(row=row, column=7, value=label).alignment = right
        ws.cell(row=row, column=7).font = bold
        c = ws.cell(row=row, column=9, value=float(value))
        c.number_format = '"₹"#,##0.00'
        c.alignment = right
        if label == "Grand Total":
            ws.cell(row=row, column=7).font = Font(bold=True, size=12, color="1F3A5F")
            ws.cell(row=row, column=9).font = Font(bold=True, size=12, color="1F3A5F")
        row += 1

    if header.get("notes"):
        row += 1
        ws.cell(row=row, column=1, value="Notes").font = bold
        row += 1
        ws.cell(row=row, column=1, value=_safe(header.get("notes")))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        ws.cell(row=row, column=1).alignment = wrap

    widths = [5, 12, 18, 42, 14, 8, 14, 9, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)
    return path
